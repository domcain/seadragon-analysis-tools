import xlrd
import xlwt
import openpyxl as xl

inat_fieldname_date = "observed_on"
sds_fieldname_year = "encounter.year"
sds_fieldname_month = "encounter.month"
sds_fieldname_day = "encounter.day"
part_of_new_excel_filename = "Seadragon Analysis"
new_excel_file_extension = ".xls"
types_of_seadragon = ["Common Seadragons", "Leafy Seadragons", "Ruby Seadragons"]


def my_string_split(text, separator):
    lst = []
    cur = ''
    currently_inside_quotes = False
    for c in text:
        if c == '"':
            currently_inside_quotes = not currently_inside_quotes
        elif c == separator and not currently_inside_quotes:
            lst.append(cur)
            cur = ''
        else:
            cur += c
    lst.append(cur)
    return lst


def format_date(year, month, day):
    if len(year) == 2:
        year = '20' + year
    if len(month) == 1:
        month = '0' + month
    if len(day) == 1:
        day = '0' + day
    return year + '-' + month + '-' + day


def analyse_data_files(sds_filename, inat_filenames):
    # Ensure that the arguments passed to 'main' are provided in the correct format

    if not isinstance(sds_filename, str):
        return [False, "The Seadragon Search filename passed to analyse_data_files was not of type 'string'"]

    if not isinstance(inat_filenames, list):
        return [False, "The list of iNaturalist filenames passed to analyse_data_files was not of type 'list'"]

    if len(inat_filenames) == 0:
        return [False, "The list of iNaturalist filenames passed to analyse_data_files was empty"]

    for inat in inat_filenames:
        if not isinstance(inat, str):
            if len(inat_filenames == 1):
                return [False, "The iNaturalist filename in the list passed to analyse_data_files was not of type 'string'"]
            else:
                return [False, "The iNaturalist filenames in the list passed to analyse_data_files were not all of type 'string'"]



    # Open the iNaturalist csv file
    inat_files = []
    for inat in inat_filenames:
        try:
            f = open(inat, "r")
        except:
            return [False, "The iNaturalist file '" + inat + "' cannot be found or cannot be opened"]
        inat_files.append(f)

    # Read in the data from the iNaturalist csv files
    inat_data = []
    each_file_inat_field_date = []
    for i in range(len(inat_files)):
        this_file = inat_files[i]

        this_file_inat_data = [] # index 0 is the headings
        for line in my_string_split(this_file.read().strip(), '\n'):
            line = my_string_split(line.lower(), ',')
            this_file_inat_data.append(line)
        this_file.close()

        # Find the position of the relevant fieldname in this iNat csv file
        try:
            this_file_inat_field_date = this_file_inat_data[0].index(inat_fieldname_date)
        except:
            return [False, "Could not find column heading '" + inat_fieldname_date + "' in iNaturalist file '" + inat_filenames[i] + "'"]

        inat_data.append(this_file_inat_data)
        each_file_inat_field_date.append(this_file_inat_field_date)

    if '.' not in sds_filename:
        return [False, "The Seadragon Search file does not have a file extension"]
    file_extension_without_dot = sds_filename.split('.')[-1]

    if file_extension_without_dot == "xls":
        # Open the Seadragon Search Excel file
        try:
            sds_wb = xlrd.open_workbook(sds_filename)
        except:
            return [False, "The Seadragon Search file cannot be found or cannot be opened as an Excel file"]
        try:
            sds_ws = sds_wb.sheet_by_index(0)
        except:
            return [False, "The Seadragon Search Excel file does not contain any worksheets"]


        # Find the minimum row and minimum column in the Seadragon Search Excel file
        sds_min_row = -1
        done = False
        for r in range(0, sds_ws.nrows):
            for c in range(0, sds_ws.ncols):
                if sds_ws.cell_value(rowx=r,colx=c) is not None:
                    sds_min_row = r
                    done = True
                    break
            if done:
                break

        sds_min_column = -1
        done = False
        for c in range(0, sds_ws.ncols):
            for r in range(0, sds_ws.nrows):
                if sds_ws.cell_value(rowx=r,colx=c) is not None:
                    sds_min_column = c
                    done = True
                    break
            if done:
                break

        if sds_min_row == -1:
            assert(sds_min_column == -1)
            return [False, "The (first worksheet in the) Seadragon Search Excel file is empty"]

        # Find the maximum row and maximum column in the Seadragon Search Excel file (worksheet.nrows-1 and worksheet.ncols-1 may be too high)
        sds_max_row = 0
        for r in range(sds_min_row, sds_ws.nrows):
            empty_row = True
            for c in range(sds_min_column, sds_ws.ncols):
                if sds_ws.cell_value(rowx=r,colx=c) is not None:
                    empty_row = False
                    break
            if empty_row:
                break
            sds_max_row = r

        sds_max_column = 0
        for c in range(sds_min_column, sds_ws.ncols):
            empty_column = True
            for r in range(sds_min_row, sds_ws.nrows):
                if sds_ws.cell_value(rowx=r,colx=c) is not None:
                    empty_column = False
                    break
            if empty_column:
                break
            sds_max_column = c


        # Read in the data from the Seadragon Search Excel file
        sds_data = [] # index 0 is the headings
        for r in range(sds_min_row, sds_max_row + 1):
            line = []
            for c in range(sds_min_column, sds_max_column + 1):
                cell_value = sds_ws.cell_value(rowx=r,colx=c)
                if isinstance(cell_value, str):
                    cell_value = cell_value.lower()
                line.append(cell_value)
            sds_data.append(line)
        assert(sds_data)

    else: # file_extension_without_dot != "xls"
        # Open the Seadragon Search Excel file
        try:
            sds_wb = xl.load_workbook(sds_filename)
        except:
            return [False, "The Seadragon Search file cannot be found or cannot be opened as an Excel file"]
        try:
            sds_ws = sds_wb.worksheets[0]
        except:
            return [False, "The Seadragon Search Excel file does not contain any worksheets"]
        
        # Find the minimum row and minimum column in the Seadragon Search Excel file
        sds_min_row = 0
        done = False
        for r in range(1, sds_ws.max_row + 1):
            for c in range(1, sds_ws.max_column + 1):
                if sds_ws.cell(row=r,column=c) is not None:
                    sds_min_row = r
                    done = True
                    break
            if done:
                break

        sds_min_column = 0
        done = False
        for c in range(1, sds_ws.max_column + 1):
            for r in range(1, sds_ws.max_row + 1):
                if sds_ws.cell(row=r,column=c) is not None:
                    sds_min_column = c
                    done = True
                    break
            if done:
                break
        
        if sds_min_row == 0:
            assert(sds_min_column == 0)
            return [False, "The (first worksheet in the) Seadragon Search Excel file is empty"]

        # Find the maximum row and maximum column in the Seadragon Search Excel file (worksheet.max_row and worksheet.max_column can be too high)
        sds_max_row = 0
        for r in range(sds_min_row, sds_ws.max_row + 1):
            empty_row = True
            for c in range(sds_min_column, sds_ws.max_column + 1):
                if sds_ws.cell(row=r,column=c) is not None:
                    empty_row = False
                    break
            if empty_row:
                break
            sds_max_row = r

        sds_max_column = 0
        for c in range(sds_min_column, sds_ws.max_column + 1):
            empty_column = True
            for r in range(sds_min_row, sds_ws.max_row + 1):
                if sds_ws.cell(row=r,column=c) is not None:
                    empty_column = False
                    break
            if empty_column:
                break
            sds_max_column = c


        # Read in the data from the Seadragon Search Excel file
        sds_data = [] # index 0 is the headings
        for r in range(sds_min_row, sds_max_row + 1):
            line = []
            for c in range(sds_min_column, sds_max_column + 1):
                cell_value = sds_ws.cell(row=r,column=c)
                if isinstance(cell_value, str):
                    cell_value = cell_value.lower()
                line.append(cell_value)
            sds_data.append(line)
        assert(sds_data)

    # Find the position of the relevant fieldnames in the Seadragon Search Excel file
    try:
        sds_field_year = sds_data[0].index(sds_fieldname_year)
    except:
        return [False, "Could not find column heading '" + sds_fieldname_year + "' in Seadragon Search file"]
    try:
        sds_field_month = sds_data[0].index(sds_fieldname_month)
    except:
        return [False, "Could not find column heading '" + sds_fieldname_month + "' in Seadragon Search file"]
    try:
        sds_field_day = sds_data[0].index(sds_fieldname_day)
    except:
        return [False, "Could not find column heading '" + sds_fieldname_day + "' in Seadragon Search file"]

    # Find the date of each iNaturalist entry
    each_file_inat_entries_on_this_day = []
    each_file_inat_rows_missing_valid_date = []
    for i in range(len(inat_files)):
        this_file_inat_data = inat_data[i]
        this_file_inat_field_date = each_file_inat_field_date[i]
        this_file_inat_entries_on_this_day = {}
        this_file_inat_rows_missing_valid_date = []
        for j in range(1, len(this_file_inat_data)):
            row = this_file_inat_data[j]
            date = str(row[this_file_inat_field_date])

            date_delimiter = "not_set_yet"
            for c in date:
                if c not in "0123456789":
                    date_delimiter = c
                    break
            if date_delimiter == "not_set_yet":
                this_file_inat_rows_missing_valid_date.append(j)
                continue
            if date.count(date_delimiter) != 2:
                this_file_inat_rows_missing_valid_date.append(j)
                continue
            year, month, day = date.split(date_delimiter)

            try:
                int(year)
                int(month)
                int(day)
            except:
                this_file_inat_rows_missing_valid_date.append(j)
                continue

            # If we can tell that the day and year are given the other way around, then swap the variables
            if len(day) == 4 and len(year) <= 2:
                day, year = year, day

            my_date_string = format_date(year, month, day)

            if my_date_string not in this_file_inat_entries_on_this_day:
                this_file_inat_entries_on_this_day[my_date_string] = []
            this_file_inat_entries_on_this_day[my_date_string].append(j)
        each_file_inat_entries_on_this_day.append(this_file_inat_entries_on_this_day)
        each_file_inat_rows_missing_valid_date.append(this_file_inat_rows_missing_valid_date)



    # Find the date of each Seadragon Search entry
    num_sds_entries_on_this_day = {}
    sds_rows_missing_date = []
    for i in range(1, len(sds_data)):
        row = sds_data[i]
        year = str(row[sds_field_year])
        month = str(row[sds_field_month])
        day = str(row[sds_field_day])
        try:
            int(year)
            int(month)
            int(day)
        except:
            sds_rows_missing_date.append(i)
            continue

        my_date_string = format_date(year, month, day)

        if my_date_string not in num_sds_entries_on_this_day:
            num_sds_entries_on_this_day[my_date_string] = 0
        num_sds_entries_on_this_day[my_date_string] += 1

    # Find the number of iNaturalist entries on each date
    num_inat_entries_on_this_day = {}
    for f in each_file_inat_entries_on_this_day:
        for date, entries in f.items():
            if date not in num_inat_entries_on_this_day:
                num_inat_entries_on_this_day[date] = 0
            num_inat_entries_on_this_day[date] += len(entries)

    # See which dates have more iNaturalist entries than Seadragon Search entries (and by how many entries)
    dates_flagged = {}
    for date, num_inat_entries in num_inat_entries_on_this_day.items():
        num_sds_entries = num_sds_entries_on_this_day.get(date, 0)
        diff = num_inat_entries - num_sds_entries
        if diff > 0:
            dates_flagged[date] = diff

    if dates_flagged:
        preview = "The Seadragon Search database appears to be missing:\n"
        for date, diff in dates_flagged.items():
            if diff == 1:
                preview += str(diff) + " iNaturalist entry from " + date + "\n"
            else:
                preview += str(diff) + " iNaturalist entries from " + date + "\n"
    else:
        preview = "The Seadragon Search database appears to be up to date"


    # Make a new Excel file for the results
    new_wb = xlwt.Workbook()

    # Display the preview in one of the worksheets of the new Excel file
    new_ws = new_wb.add_sheet("Preview")
    row = 0
    for line in preview.split("\n"):
        new_ws.write(row, 0, line)
        row += 1


    style = xlwt.easyxf("pattern: pattern solid, fore_colour yellow")

    for i in range(len(inat_files)):

        # Create a new worksheet in the new Excel file
        sheet_name = "Sheet" + str(i + 2) # + 1 because the sheets should be 1-indexed, and another + 1 because the first sheet contains the preview
        for type in types_of_seadragon:
            if type.lower() in inat_filenames[i].lower():
                sheet_name = type
                break
        new_ws = new_wb.add_sheet(sheet_name)
        this_file_inat_data = inat_data[i]

        # Store precisely which rows should be highlighted in this worksheet of the new Excel file
        should_highlight_row = []
        for j in range(len(this_file_inat_data)):
            should_highlight_row.append(False)
        for date in dates_flagged:
            if date in each_file_inat_entries_on_this_day[i]:
                for row in each_file_inat_entries_on_this_day[i][date]:
                    should_highlight_row[row] = True

        # Fill this worksheet of the new Excel file with data from the corresponding iNaturalist csv file, highlighting rows appropriately
        for r in range(len(this_file_inat_data)):
            for c in range(len(this_file_inat_data[0])):
                if should_highlight_row[r]:
                    new_ws.write(r, c, this_file_inat_data[r][c], style)
                else:
                    new_ws.write(r, c, this_file_inat_data[r][c])

    # Choose a name for the new Excel file
    name = "Unnamed"
    breaking = False
    for filename in inat_filenames:
        for type in types_of_seadragon:
            if type.lower() in filename.lower():
                name_in_filename = filename[:filename.lower().index(type.lower())].strip()
                if name != "Unnamed":
                    # If this iNaturalist filename has a different person's name to one of the other iNaturalist filenames,
                    # then the person's name cannot be interpreted
                    if name_in_filename.lower() != name.lower():
                        name = "Unnamed"
                        breaking = True
                        break
                else:
                    name = name_in_filename
        if breaking:
            break


    suggested_filename = name + " " + part_of_new_excel_filename + new_excel_file_extension
    return [True, preview, suggested_filename, new_wb]





    # DEBUG
    print("\nDaily iNaturalist entries:")
    print(num_inat_entries_on_this_day)
    print("Daily Seadragon Search entries:")
    print(num_sds_entries_on_this_day)